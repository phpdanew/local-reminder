import json
import threading
from datetime import datetime, timedelta
from pync import Notifier
from abc import ABC, abstractmethod
import openpyxl
import os
from openpyxl.styles import Font

class ReminderWriter(ABC):
    """提醒写入器抽象基类"""
    
    @abstractmethod
    def write_reminder(self, text, delay_minutes, reminder_id=None):
        """写入提醒信息"""
        pass
    
    @abstractmethod
    def update_completion(self, text, reminder_id=None):
        """更新完成时间"""
        pass

class LogWriter(ReminderWriter):
    """日志文件写入器"""
    
    def __init__(self, log_file='reminders.log'):
        self.log_file = log_file
    
    def write_reminder(self, text, delay_minutes, reminder_id=None):
        """写入提醒信息到日志"""
        start_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        end_time = (datetime.now() + timedelta(minutes=delay_minutes)).strftime('%Y-%m-%d %H:%M:%S')
        log_entry = f"{start_time} {start_time} {text}\n"
        
        with open(self.log_file, 'a', encoding='utf-8') as f:
            f.write(log_entry)
        
        return start_time, end_time
    
    def update_completion(self, text, reminder_id=None):
        """更新最后一行的结束时间"""
        completion_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # 读取所有行
        try:
            with open(self.log_file, 'r', encoding='utf-8') as f:
                lines = f.readlines()
        except FileNotFoundError:
            return
        
        if not lines:
            return
        
        # 更新最后一行
        last_line = lines[-1].strip()
        if last_line:
            # 分割行内容
            parts = last_line.split(' ', 2)  # 最多分割2次，保留任务文本的完整性
            if len(parts) >= 3:
                start_time, _, task_text = parts
                # 更新结束时间
                lines[-1] = f"{start_time} {completion_time} {task_text}\n"
        
        # 写回文件
        with open(self.log_file, 'w', encoding='utf-8') as f:
            f.writelines(lines)

class ExcelWriter(ReminderWriter):
    """Excel文件写入器"""
    
    def __init__(self, excel_file='plan.xlsx'):
        self.excel_file = excel_file
        self.sheet_name = '当前'
        self._ensure_file_exists()
    
    def _ensure_file_exists(self):
        """确保Excel文件存在，如果不存在则创建"""
        if not os.path.exists(self.excel_file):
            wb = openpyxl.Workbook()
            
            # 创建三个sheet
            wb.create_sheet('当前', 0)
            wb.create_sheet('历史', 1)
            wb.create_sheet('分析', 2)
            
            # 删除默认的Sheet
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            
            # 设置表头
            header_font = Font(name='宋体', size=16)
            
            for sheet_name in ['当前', '历史', '分析']:
                ws = wb[sheet_name]
                ws['A1'] = '开始时间'
                ws['B1'] = '结束时间'
                ws['C1'] = '任务'
                ws['D1'] = '延时分钟'
                
                # 设置表头字体
                ws['A1'].font = header_font
                ws['B1'].font = header_font
                ws['C1'].font = header_font
                ws['D1'].font = header_font
            
            wb.save(self.excel_file)
            wb.close()
    
    def _get_next_row(self, sheet):
        """获取下一个空行"""
        for row in range(1, sheet.max_row + 2):
            if not sheet[f'A{row}'].value:
                return row
        return sheet.max_row + 1
    
    def write_reminder(self, text, delay_minutes, reminder_id=None):
        """写入提醒信息到Excel"""
        wb = openpyxl.load_workbook(self.excel_file)
        ws = wb[self.sheet_name]
        
        start_time = datetime.now()
        end_time = start_time + timedelta(minutes=delay_minutes)
        
        # 获取下一个空行
        next_row = self._get_next_row(ws)
        
        # 写入数据
        ws[f'A{next_row}'] = start_time.strftime('%Y-%m-%d %H:%M:%S')
        ws[f'C{next_row}'] = text
        ws[f'D{next_row}'] = delay_minutes
        
        font = Font(name='宋体', size=16)
        ws[f'A{next_row}'].font = font
        ws[f'C{next_row}'].font = font
        ws[f'D{next_row}'].font = font
        
        wb.save(self.excel_file)
        wb.close()
        
        return start_time.strftime('%Y-%m-%d %H:%M:%S'), end_time.strftime('%Y-%m-%d %H:%M:%S')
    
    def update_completion(self, text, reminder_id=None):
        """更新完成时间到Excel"""
        wb = openpyxl.load_workbook(self.excel_file)
        ws = wb[self.sheet_name]
        
        completion_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # 查找最后一行有数据的行并更新结束时间
        last_row = ws.max_row
        while last_row > 1 and not ws[f'A{last_row}'].value:
            last_row -= 1
        
        if last_row > 1:  # 确保不是表头行
            font = Font(name='宋体', size=16)
            ws[f'B{last_row}'] = completion_time  # 只更新B列（结束时间）
            ws[f'B{last_row}'].font = font
        
        wb.save(self.excel_file)
        wb.close()

class ReminderLogic:
    def __init__(self, writer=None):
        # 全局变量存储定时任务
        self.current_timer = None
        self.timer_lock = threading.Lock()
        
        # 使用传入的写入器，默认使用日志写入器
        self.writer = writer if writer else ExcelWriter()
    
    def process_reminder_request(self, data):
        """处理提醒请求"""
        try:
            if not data or 'text' not in data:
                return {
                    'success': False,
                    'error': '缺少文本内容',
                    'status_code': 400
                }
            
            text = data['text']
            timestamp = data.get('timestamp', datetime.now().isoformat())
            
            # 获取延时时间，默认30分钟
            delay_minutes = data.get('delay_minutes', 30)
            
            # 先写入文件
            start_time, end_time = self.writer.write_reminder(text, delay_minutes)
            
            print(f"已记录提醒: {text} (延时{delay_minutes}分钟)")
            
            # 设置延时通知
            self._set_delayed_notification(text, timestamp, delay_minutes)
            
            # 返回成功响应
            return {
                'success': True,
                'message': f'提醒已设置，{delay_minutes}分钟后通知',
                'text': text,
                'timestamp': timestamp,
                'delay_minutes': delay_minutes,
                'start_time': start_time,
                'end_time': end_time
            }
            
        except Exception as e:
            print(f"处理提醒请求时出错: {str(e)}")
            return {
                'success': False,
                'error': f'服务器错误: {str(e)}',
                'status_code': 500
            }
    
    def _set_delayed_notification(self, text, timestamp, delay_minutes):
        """设置延时通知，确保只有一个定时任务"""
        with self.timer_lock:
            # 取消之前的定时任务
            if self.current_timer:
                self.current_timer.cancel()
                print("已取消之前的定时任务")
            
            # 创建新的定时任务
            self.current_timer = threading.Timer(
                delay_minutes * 60, 
                self._send_notification, 
                args=[text, timestamp]
            )
            self.current_timer.daemon = True
            self.current_timer.start()
            
            # 计算通知时间
            notification_time = datetime.now() + timedelta(minutes=delay_minutes)
            print(f"已设置新的定时任务: {text}")
            print(f"通知时间: {notification_time.strftime('%Y-%m-%d %H:%M:%S')}")
    
    def _send_notification(self, text, timestamp):
        """发送系统通知"""
        try:
            # 使用pync发送系统通知
            Notifier.notify(
                text,
                title='本地提醒器',
                subtitle=f'时间: {timestamp}',
                sound='default'
            )
            
            print(f"已发送延时通知: {text}")
            
            # 更新完成时间
            self.writer.update_completion(text)
                
        except Exception as e:
            print(f"发送通知时出错: {str(e)}")
    
    def get_status(self):
        """获取当前定时任务状态"""
        with self.timer_lock:
            if self.current_timer and self.current_timer.is_alive():
                return {
                    'status': 'active',
                    'message': '有定时任务正在运行',
                    'timestamp': datetime.now().isoformat()
                }
            else:
                return {
                    'status': 'idle',
                    'message': '没有定时任务',
                    'timestamp': datetime.now().isoformat()
                }
    
    def cleanup(self):
        """清理定时任务"""
        if self.current_timer:
            self.current_timer.cancel()
            print("已清理定时任务")
